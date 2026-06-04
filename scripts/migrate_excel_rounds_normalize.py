#!/usr/bin/env python3
"""
엑셀/RELEASE-* 마이그레이션 데이터 → 타임라인용 12개 라운드 + 구성 + RC1 통일 정규화
================================================================================
실행 순서 (프로젝트 루트):

  python scripts/migrate_excel_to_db.py [excel.xlsx]
  python scripts/migrate_excel_rounds_normalize.py [excel.xlsx]        # dry-run
  python scripts/migrate_excel_rounds_normalize.py [excel.xlsx] --apply

원칙:
  · 최상위 12 라운드: upstream=MULTI_PRODUCT, release_visible=1, remark에 수행 기간
  · 구성(토폴로지)마다 RC1 하나만 (RC2+ Run은 RC1으로 합침)
  · product_test_run.product_test_release_id → 해당 구성의 RC1
  · 기존 RELEASE-* 평면 release는 round_legacy 로 숨김 (삭제하지 않음)

선택 인자:
  --apply     실제 DB 반영 (없으면 dry-run)
  --no-excel  DB/remark 만으로 기간·매핑 (엑셀 생략)
"""
from __future__ import annotations

import re
import shutil
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
DB_PATH = PROJECT_ROOT / "data" / "product_test_tracking_system.db"
DEFAULT_EXCEL = PROJECT_ROOT.parent / "ai_coworking" / "test_tracing_system_to_migrate.xlsx"
PROJECT_ID = "WIFI_CONNECTIVITY_TEST_2026"
MULTI_PRODUCT = "MULTI_PRODUCT"
ACTOR = "excel_rounds_normalize_v1"
NOW = datetime.now(timezone.utc).isoformat()
RC_SEQ = 1

# ── 타임라인 최상위 12 라운드 (사용자 정의) ─────────────────────────────────────
# patterns: rid/alias/remark 등에 매칭 (긴/구체적 라운드가 먼저 오도록 정렬해 사용)
CANONICAL_ROUNDS: list[dict] = [
    {
        "short": "WIFI_1ST_IMPROVE",
        "seq": 2,
        "alias": "Wi-Fi Connectivity Test 1차 개선확인 시험 (결함 제품)",
        "patterns": [
            r"WIFI[_\s.-]*1[_\s.-]*ST[_\s.-]*IMPROVE",
            r"1차.*개선",
            r"1ST.*IMPROVE",
        ],
    },
    {
        "short": "WIFI_2ND_IMPROVE",
        "seq": 4,
        "alias": "Wi-Fi Connectivity Test 2차 개선확인 시험 (결함 제품)",
        "patterns": [
            r"WIFI[_\s.-]*2[_\s.-]*ND[_\s.-]*IMPROVE",
            r"2차.*개선",
            r"2ND.*IMPROVE",
        ],
    },
    {
        "short": "WIFI_1ST",
        "seq": 1,
        "alias": "Wi-Fi Connectivity Test 1차 (5개 제품)",
        "patterns": [
            r"WIFI[_\s.-]*1[_\s.-]*ST(?!.*IMPROVE)",
            r"CONNECTIVITY.*1차",
            r"1차.*WIFI",
        ],
    },
    {
        "short": "WIFI_2ND",
        "seq": 3,
        "alias": "Wi-Fi Connectivity Test 2차 (5개 제품)",
        "patterns": [
            r"WIFI[_\s.-]*2[_\s.-]*ND(?!.*IMPROVE)",
            r"CONNECTIVITY.*2차",
            r"2차.*WIFI",
        ],
    },
    {
        "short": "HDC_9100_1_0_5A",
        "seq": 5,
        "alias": "HDC-9100 1.0.5A 시험",
        "patterns": [r"HDC[_\s.-]*9100[_\s.-]*1[_\s.-]*0[_\s.-]*5A", r"HDC_1_0_5A"],
    },
    {
        "short": "HDR_9000_1_1_7E",
        "seq": 6,
        "alias": "HDR-9000 1.1.7E 시험",
        "patterns": [r"HDR[_\s.-]*9000[_\s.-]*1[_\s.-]*1[_\s.-]*7E"],
    },
    {
        "short": "HDR_9000_1_1_8",
        "seq": 7,
        "alias": "HDR-9000 1.1.8 시험",
        "patterns": [r"HDR[_\s.-]*9000[_\s.-]*1[_\s.-]*1[_\s.-]*8(?!.*7E)"],
    },
    {
        "short": "HLM_9000_1_1_14B",
        "seq": 8,
        "alias": "HLM-9000 1.1.14B 시험",
        "patterns": [r"HLM[_\s.-]*9000[_\s.-]*1[_\s.-]*1[_\s.-]*14B"],
    },
    {
        "short": "HRK_9000A_1_1_0A",
        "seq": 9,
        "alias": "HRK-9000A 1.1.0A 시험",
        "patterns": [r"HRK[_\s.-]*9000A[_\s.-]*1[_\s.-]*1[_\s.-]*0A", r"DOWNGRADE"],
    },
    {
        "short": "HRK_9000A_1_1_1D",
        "seq": 11,
        "alias": "HRK-9000A 1.1.1D 시험",
        "patterns": [r"HRK[_\s.-]*9000A[_\s.-]*1[_\s.-]*1[_\s.-]*1D", r"WIFI_1_1_1D"],
    },
    {
        "short": "HRK_9000A_1_1_1A",
        "seq": 10,
        "alias": "HRK-9000A 1.1.1A 시험",
        "patterns": [r"HRK[_\s.-]*9000A[_\s.-]*1[_\s.-]*1[_\s.-]*1A(?!.*1D)"],
    },
    {
        "short": "HTR_1A_1_1_8",
        "seq": 12,
        "alias": "HTR-1A 1.1.8 시험",
        "patterns": [r"HTR[_\s.-]*1A[_\s.-]*1[_\s.-]*1[_\s.-]*8"],
    },
]

CANONICAL_SHORTS = {r["short"] for r in CANONICAL_ROUNDS}
CANONICAL_BY_SHORT = {r["short"]: r for r in CANONICAL_ROUNDS}
# prefix 매칭: 긴 short 우선
CANONICAL_SHORTS_SORTED = sorted(CANONICAL_SHORTS, key=len, reverse=True)

DEVICE_ORDER = ["HRK", "HTR", "HLM", "HDR", "HDC", "HIIS"]
STATUS_PRIORITY = {
    "BLOCKED": 0,
    "TESTING": 1,
    "DRAFT": 2,
    "PASSED": 3,
    "QI_TEAM_RELEASED": 3,
    "APPROVED": 3,
    "QI_TEAM_REVIEWED": 4,
    "DONE": 5,
}


def normalize_combo(raw: str) -> str | None:
    if not raw or raw.strip() in ("TBD", "VARIOUS_CONNECTIONS", ""):
        return None
    combo = raw.strip().replace(" ", "")
    parts = re.findall(r"(\d*)(AP|HRK|HTR|HLM|HDR|HDC|HIIS)", combo)
    if not parts:
        return None
    ap_part = ""
    device_counts: dict[str, int] = {}
    for count_str, device in parts:
        count = int(count_str) if count_str else 1
        if device == "AP":
            ap_part = f"{count}AP"
        else:
            device_counts[device] = device_counts.get(device, 0) + count
    if not ap_part:
        return None
    sorted_devices = sorted(
        device_counts.items(),
        key=lambda x: DEVICE_ORDER.index(x[0]) if x[0] in DEVICE_ORDER else 99,
    )
    device_part = "_".join(f"{cnt}{dev}" for dev, cnt in sorted_devices)
    return f"{ap_part}_{device_part}" if device_part else None


def extract_combo_from_remark(remark: str) -> str:
    match = re.search(r"\[연결구성\]\s*(.+)", remark or "")
    return match.group(1).strip() if match else ""


def extract_topo_from_case_id(case_id: str) -> str:
    match = re.match(r"TEST_CASE-([^-]+)-", case_id or "")
    return match.group(1) if match else ""


def infer_round_short(text: str) -> str | None:
    blob = (text or "").strip()
    if not blob:
        return None
    normalized = blob.upper().replace(" ", "_")
    for round_def in CANONICAL_ROUNDS:
        for pattern in round_def["patterns"]:
            if re.search(pattern, normalized, re.IGNORECASE):
                return round_def["short"]
    return None


def release_id_to_round_short(release_id: str, rid_map: dict[str, str]) -> str | None:
    if not release_id:
        return None
    if release_id.startswith("RELEASE-"):
        key = release_id[len("RELEASE-") :]
        return rid_map.get(key) or rid_map.get(release_id) or infer_round_short(key)

    if release_id.startswith("TEST_RELEASE-"):
        body = release_id[len("TEST_RELEASE-") :]
        for short in CANONICAL_SHORTS_SORTED:
            if body == short or body.startswith(short + "-"):
                return short
        return infer_round_short(body)

    return infer_round_short(release_id)


def walk_round_short(conn: sqlite3.Connection, release_id: str, rid_map: dict[str, str]) -> str | None:
    cur = release_id
    for _ in range(14):
        if not cur:
            break
        short = release_id_to_round_short(cur, rid_map)
        if short:
            return short
        row = conn.execute(
            "SELECT upstream_release_id FROM product_test_release WHERE product_test_release_id=?",
            (cur,),
        ).fetchone()
        if not row or not row[0]:
            break
        cur = row[0]
        if cur == MULTI_PRODUCT:
            break
    return None


def format_period_remark(alias: str, workday: str, start: str, end: str) -> str:
    lines = [f"[Report Alias] {alias}"]
    if workday:
        lines.append(f"[Workday] {workday}")
    start_s = (start or "").strip()
    end_s = (end or "").strip()
    if start_s or end_s:
        lines.append(f"[Start] {start_s}  [End] {end_s}")
    return "\n".join(lines)


def merge_period(
    existing: dict[str, str],
    workday: str | None,
    start: str | None,
    end: str | None,
) -> dict[str, str]:
    out = dict(existing)
    if workday and str(workday).strip():
        out["workday"] = str(workday).strip()
    start_s = (start or "").strip()
    end_s = (end or "").strip()
    if start_s:
        if not out.get("start") or start_s < out["start"]:
            out["start"] = start_s
    if end_s:
        if not out.get("end") or end_s > out["end"]:
            out["end"] = end_s
    return out


def load_excel_reports(
    excel_path: Path,
) -> tuple[list[dict], dict[str, str], dict[str, dict[str, str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Reports"] if "Reports" in wb.sheetnames else wb[wb.sheetnames[0]]

    def safe(v) -> str:
        return "" if v is None else str(v).strip()

    reports: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        reports.append(
            {
                "alias": safe(row[1]),
                "rids_raw": safe(row[5]),
                "workday": safe(row[6]) or None,
                "start": safe(row[7]) or None,
                "end": safe(row[8]) or None,
            }
        )
    wb.close()

    rid_map: dict[str, str] = {}
    period_by_short: dict[str, dict[str, str]] = defaultdict(
        lambda: {"workday": "", "start": "", "end": ""}
    )

    for rpt in reports:
        candidates: list[str] = []
        if rpt["rids_raw"]:
            for part in re.split(r"[\n\r,;]+", rpt["rids_raw"]):
                part = part.strip()
                if part:
                    candidates.append(part)
        if rpt["alias"]:
            candidates.append(rpt["alias"])

        round_short = None
        for cand in candidates:
            round_short = infer_round_short(cand)
            if round_short:
                break

        if not round_short:
            continue

        for cand in candidates:
            rid_map[cand] = round_short
            rid_map[f"RELEASE-{cand}"] = round_short

        period_by_short[round_short] = merge_period(
            period_by_short[round_short],
            rpt["workday"],
            rpt["start"],
            rpt["end"],
        )

    return reports, rid_map, period_by_short


def ensure_release_visible_column(conn: sqlite3.Connection) -> None:
    cols = {row[1] for row in conn.execute("PRAGMA table_info(product_test_release)")}
    if "release_visible" not in cols:
        conn.execute(
            "ALTER TABLE product_test_release ADD COLUMN release_visible INTEGER NOT NULL DEFAULT 1"
        )


def calc_release_status_from_results(
    conn: sqlite3.Connection, release_id: str
) -> str:
    rows = conn.execute(
        """
        SELECT res.product_test_result_status, COUNT(*)
        FROM product_test_result res
        JOIN product_test_run run ON run.product_test_run_id = res.product_test_run_id
        WHERE run.product_test_release_id = ?
        GROUP BY res.product_test_result_status
        """,
        (release_id,),
    ).fetchall()
    if not rows:
        return "TESTING"
    status_map = dict(rows)
    blocked = status_map.get("blocked", 0)
    testing = status_map.get("testing", 0)
    passed = status_map.get("passed", 0)
    open_defects = conn.execute(
        """
        SELECT COUNT(*) FROM product_test_defect def
        JOIN product_test_result res ON res.product_test_result_id = def.product_test_result_id
        JOIN product_test_run run ON run.product_test_run_id = res.product_test_run_id
        WHERE run.product_test_release_id = ?
          AND def.product_test_defect_status = 'opened'
        """,
        (release_id,),
    ).fetchone()[0]
    if blocked > 0 or open_defects > 0:
        return "BLOCKED"
    if testing > 0:
        return "TESTING"
    if passed > 0:
        return "PASSED"
    return "TESTING"


def main() -> None:
    apply_mode = "--apply" in sys.argv
    no_excel = "--no-excel" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    excel_path = Path(args[0]) if args else DEFAULT_EXCEL

    print("=" * 72)
    print(f"  엑셀 라운드 정규화 (RC1 통일) [{'적용' if apply_mode else 'DRY-RUN'}]")
    print(f"  DB    : {DB_PATH}")
    print(f"  Excel : {excel_path if not no_excel else '(생략)'}")
    print("=" * 72)

    if not DB_PATH.exists():
        sys.exit(f"[ERROR] DB 없음: {DB_PATH}")

    rid_map: dict[str, str] = {}
    period_by_short: dict[str, dict[str, str]] = defaultdict(
        lambda: {"workday": "", "start": "", "end": ""}
    )

    if not no_excel and excel_path.exists():
        try:
            reports, rid_map, period_by_short = load_excel_reports(excel_path)
            print(f"[엑셀] Reports {len(reports)}행, rid 매핑 {len(rid_map)}건")
        except Exception as exc:
            print(f"[WARN] 엑셀 로드 실패 ({exc}) — DB 기반만 진행")
    elif not no_excel:
        print(f"[WARN] 엑셀 없음: {excel_path} — DB/remark 기반만 진행")

    backup_path = DB_PATH.with_suffix(
        f".backup_normalize_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    )
    if apply_mode:
        shutil.copy2(DB_PATH, backup_path)
        print(f"[백업] {backup_path}")

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=OFF")
    ensure_release_visible_column(conn)

    before_results = conn.execute("SELECT COUNT(*) FROM product_test_result").fetchone()[0]

    # DB 기존 RELEASE remark 에서 기간 보강
    for row in conn.execute(
        """
        SELECT product_test_release_id, upstream_release_id, remark
        FROM product_test_release
        WHERE product_test_release_id LIKE 'RELEASE-%'
           OR upstream_release_system = 'EXCEL_MIGRATION'
        """
    ):
        rel_id = row[0]
        upstream = row[1] or ""
        remark = row[2] or ""
        short = rid_map.get(upstream) or rid_map.get(rel_id) or infer_round_short(upstream or rel_id)
        if not short:
            continue
        workday = start = end = ""
        for line in remark.split("\n"):
            line = line.strip()
            if line.startswith("[Workday]"):
                workday = line.replace("[Workday]", "").strip()
            elif line.startswith("[Start]"):
                rest = line.replace("[Start]", "").strip()
                if "[End]" in rest:
                    parts = rest.split("[End]", 1)
                    start = parts[0].strip()
                    end = parts[1].strip()
                else:
                    start = rest
            elif line.startswith("[End]"):
                end = line.replace("[End]", "").strip()
        period_by_short[short] = merge_period(period_by_short[short], workday, start, end)

    # ── 1) 12 라운드 UPSERT ───────────────────────────────────────────────────
    round_ids: dict[str, str] = {}
    rounds_created = 0
    for spec in CANONICAL_ROUNDS:
        short = spec["short"]
        round_id = f"TEST_RELEASE-{short}"
        round_ids[short] = round_id
        period = period_by_short.get(short, {})
        remark = format_period_remark(
            spec["alias"],
            period.get("workday", ""),
            period.get("start", ""),
            period.get("end", ""),
        )
        exists = conn.execute(
            "SELECT 1 FROM product_test_release WHERE product_test_release_id=?",
            (round_id,),
        ).fetchone()
        if apply_mode:
            if exists:
                conn.execute(
                    """
                    UPDATE product_test_release SET
                        upstream_release_id=?,
                        upstream_release_system=?,
                        release_stage=?,
                        release_sequence=?,
                        release_visible=1,
                        remark=?,
                        updated_at=?, updated_by=?
                    WHERE product_test_release_id=?
                    """,
                    (
                        MULTI_PRODUCT,
                        "ROUND_NORMALIZE",
                        "round",
                        spec["seq"],
                        remark,
                        NOW,
                        ACTOR,
                        round_id,
                    ),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO product_test_release
                      (product_test_release_id, project_id,
                       upstream_release_id, upstream_release_system,
                       release_stage, release_sequence,
                       product_test_release_status, release_visible,
                       created_at, created_by, updated_at, updated_by, remark)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        round_id,
                        PROJECT_ID,
                        MULTI_PRODUCT,
                        "ROUND_NORMALIZE",
                        "round",
                        spec["seq"],
                        "TESTING",
                        1,
                        NOW,
                        ACTOR,
                        NOW,
                        ACTOR,
                        remark,
                    ),
                )
                rounds_created += 1
        print(
            f"  라운드 {round_id}  seq={spec['seq']}  "
            f"기간={period.get('start', '-')}/{period.get('end', '-')}"
        )

    # ── 2) result → (round_short, combo) ─────────────────────────────────────
    results_raw = conn.execute(
        """
        SELECT
            res.product_test_result_id,
            res.product_test_case_id,
            res.remark,
            run.product_test_run_id,
            run.product_test_release_id,
            run.product_test_target_id,
            run.product_test_environment_id,
            run.started_at,
            run.started_by,
            run.finished_at,
            run.remark AS run_remark,
            run.product_test_run_status
        FROM product_test_result res
        JOIN product_test_run run ON run.product_test_run_id = res.product_test_run_id
        """
    ).fetchall()

    topo_meta: dict[tuple[str, str], dict] = {}
    rc_meta: dict[tuple[str, str], dict] = {}
    run_targets: dict[str, str] = {}
    unclassified = 0

    for row in results_raw:
        rc_id = row["product_test_release_id"]
        round_short = walk_round_short(conn, rc_id, rid_map)
        if not round_short:
            round_short = "UNCLASSIFIED"
            unclassified += 1

        raw_combo = extract_combo_from_remark(row["remark"] or "")
        if not raw_combo:
            raw_combo = extract_topo_from_case_id(row["product_test_case_id"] or "")
        combo = normalize_combo(raw_combo) or "UNCLASSIFIED"

        round_id = round_ids.get(round_short) or f"TEST_RELEASE-{round_short}"
        if round_short == "UNCLASSIFIED" and round_id not in round_ids.values():
            round_id = "TEST_RELEASE-UNCLASSIFIED"
            if apply_mode and not conn.execute(
                "SELECT 1 FROM product_test_release WHERE product_test_release_id=?",
                (round_id,),
            ).fetchone():
                conn.execute(
                    """
                    INSERT INTO product_test_release
                      (product_test_release_id, project_id,
                       upstream_release_id, upstream_release_system,
                       release_stage, release_sequence,
                       product_test_release_status, release_visible,
                       created_at, created_by, updated_at, updated_by, remark)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        round_id,
                        PROJECT_ID,
                        MULTI_PRODUCT,
                        "ROUND_NORMALIZE",
                        "round",
                        99,
                        "TESTING",
                        1,
                        NOW,
                        ACTOR,
                        NOW,
                        ACTOR,
                        "[Report Alias] 미분류 마이그레이션 데이터",
                    ),
                )

        topo_key = (round_short, combo)
        if topo_key not in topo_meta:
            topo_id = f"TEST_RELEASE-{round_short}-{combo}"
            period = period_by_short.get(round_short, {})
            topo_meta[topo_key] = {
                "id": topo_id,
                "round_id": round_id,
                "round_short": round_short,
                "combo": combo,
                "remark": format_period_remark(
                    combo,
                    period.get("workday", ""),
                    period.get("start", ""),
                    period.get("end", ""),
                ),
            }

        topo_id = topo_meta[topo_key]["id"]
        rc_key = (round_short, combo)
        if rc_key not in rc_meta:
            rc_id_new = f"{topo_id}-RC{RC_SEQ}"
            rc_meta[rc_key] = {
                "id": rc_id_new,
                "topo_id": topo_id,
                "round_id": round_id,
            }

        run_targets[row["product_test_run_id"]] = rc_meta[rc_key]["id"]

    print(f"\n[매핑] result {len(results_raw)}건 → 라운드 {len(topo_meta)} 구성, RC1 {len(rc_meta)}개")
    if unclassified:
        print(f"  [주의] 라운드 미식별 {unclassified}건 → UNCLASSIFIED")

    if not apply_mode:
        print("\nDRY-RUN 완료. 적용: python scripts/migrate_excel_rounds_normalize.py --apply")
        conn.close()
        return

    # ── 3) topology + RC1 생성 ───────────────────────────────────────────────
    topo_seq: dict[str, int] = defaultdict(int)
    for (round_short, _combo), info in sorted(topo_meta.items(), key=lambda x: (x[0][0], x[0][1])):
        topo_seq[round_short] += 1
        seq = topo_seq[round_short]
        if not conn.execute(
            "SELECT 1 FROM product_test_release WHERE product_test_release_id=?",
            (info["id"],),
        ).fetchone():
            conn.execute(
                """
                INSERT INTO product_test_release
                  (product_test_release_id, project_id,
                   upstream_release_id, upstream_release_system,
                   release_stage, release_sequence,
                   product_test_release_status, release_visible,
                   created_at, created_by, updated_at, updated_by, remark)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    info["id"],
                    PROJECT_ID,
                    info["round_id"],
                    "ROUND_NORMALIZE",
                    "RC",
                    seq,
                    "TESTING",
                    1,
                    NOW,
                    ACTOR,
                    NOW,
                    ACTOR,
                    info["remark"],
                ),
            )

    for _key, rc in rc_meta.items():
        if not conn.execute(
            "SELECT 1 FROM product_test_release WHERE product_test_release_id=?",
            (rc["id"],),
        ).fetchone():
            conn.execute(
                """
                INSERT INTO product_test_release
                  (product_test_release_id, project_id,
                   upstream_release_id, upstream_release_system,
                   release_stage, release_sequence,
                   product_test_release_status, release_visible,
                   created_at, created_by, updated_at, updated_by, remark)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    rc["id"],
                    PROJECT_ID,
                    rc["topo_id"],
                    "ROUND_NORMALIZE",
                    "RC",
                    RC_SEQ,
                    "TESTING",
                    0,
                    NOW,
                    ACTOR,
                    NOW,
                    ACTOR,
                    f"[구성] {_key[1]}\n[RC] RC{RC_SEQ} 통일",
                ),
            )

    # ── 4) Run → RC1, report → 라운드 ────────────────────────────────────────
    runs_updated = 0
    for run_id, target_rc in run_targets.items():
        conn.execute(
            """
            UPDATE product_test_run
            SET product_test_release_id=?, updated_at=?, updated_by=?
            WHERE product_test_run_id=?
            """,
            (target_rc, NOW, ACTOR, run_id),
        )
        runs_updated += 1

    for row in conn.execute(
        "SELECT product_test_report_id, product_test_release_id FROM product_test_report"
    ):
        short = walk_round_short(conn, row[1], rid_map)
        if short and short in round_ids:
            conn.execute(
                """
                UPDATE product_test_report
                SET product_test_release_id=?, updated_at=?, updated_by=?
                WHERE product_test_report_id=?
                """,
                (round_ids[short], NOW, ACTOR, row[0]),
            )

    # ── 5) RC2+ → RC1 합치기 ─────────────────────────────────────────────────
    merged_runs = 0
    deleted_rc = 0
    for rc_row in conn.execute(
        """
        SELECT product_test_release_id, upstream_release_id
        FROM product_test_release
        WHERE product_test_release_id LIKE '%-RC%'
          AND product_test_release_id NOT LIKE '%-RC1'
          AND COALESCE(release_visible, 1) = 0
        """
    ):
        old_rc = rc_row[0]
        parent_topo = rc_row[1]
        if not parent_topo:
            continue
        target_rc = f"{parent_topo}-RC{RC_SEQ}"
        if not conn.execute(
            "SELECT 1 FROM product_test_release WHERE product_test_release_id=?",
            (target_rc,),
        ).fetchone():
            continue
        cur = conn.execute(
            "SELECT COUNT(*) FROM product_test_run WHERE product_test_release_id=?",
            (old_rc,),
        ).fetchone()[0]
        if cur:
            conn.execute(
                """
                UPDATE product_test_run
                SET product_test_release_id=?, updated_at=?, updated_by=?
                WHERE product_test_release_id=?
                """,
                (target_rc, NOW, ACTOR, old_rc),
            )
            merged_runs += cur
        conn.execute(
            "DELETE FROM product_test_release WHERE product_test_release_id=?",
            (old_rc,),
        )
        deleted_rc += 1

    # ── 6) 평면 RELEASE-* 숨김 ────────────────────────────────────────────────
    legacy = conn.execute(
        """
        UPDATE product_test_release
        SET release_stage='round_legacy', release_visible=0,
            updated_at=?, updated_by=?
        WHERE product_test_release_id LIKE 'RELEASE-%'
        """,
        (NOW, ACTOR),
    ).rowcount

    # ── 7) 상태 보정 (RC → topology → round) ─────────────────────────────────
    all_rc = [r["id"] for r in rc_meta.values()]
    for rc_id in all_rc:
        status = calc_release_status_from_results(conn, rc_id)
        conn.execute(
            "UPDATE product_test_release SET product_test_release_status=? WHERE product_test_release_id=?",
            (status, rc_id),
        )

    for info in topo_meta.values():
        child_statuses = [
            s[0]
            for s in conn.execute(
                """
                SELECT product_test_release_status FROM product_test_release
                WHERE upstream_release_id=? AND COALESCE(release_visible,0)=0
                """,
                (info["id"],),
            ).fetchall()
        ]
        if child_statuses:
            best = min(child_statuses, key=lambda s: STATUS_PRIORITY.get(s, 99))
            conn.execute(
                "UPDATE product_test_release SET product_test_release_status=? WHERE product_test_release_id=?",
                (best, info["id"]),
            )

    for spec in CANONICAL_ROUNDS:
        round_id = round_ids[spec["short"]]
        child_statuses = [
            s[0]
            for s in conn.execute(
                """
                SELECT product_test_release_status FROM product_test_release
                WHERE upstream_release_id=? AND COALESCE(release_visible,1)=1
                """,
                (round_id,),
            ).fetchall()
        ]
        if child_statuses:
            best = min(child_statuses, key=lambda s: STATUS_PRIORITY.get(s, 99))
            conn.execute(
                "UPDATE product_test_release SET product_test_release_status=? WHERE product_test_release_id=?",
                (best, round_id),
            )

    conn.commit()

    after_results = conn.execute("SELECT COUNT(*) FROM product_test_result").fetchone()[0]
    top_levels = conn.execute(
        """
        SELECT product_test_release_id, release_sequence,
               substr(remark, 1, 80) AS remark_head
        FROM product_test_release
        WHERE upstream_release_id=?
        ORDER BY release_sequence
        """,
        (MULTI_PRODUCT,),
    ).fetchall()

    print(f"\n[적용 완료]")
    print(f"  신규 라운드: {rounds_created}")
    print(f"  Run→RC1: {runs_updated}건")
    print(f"  RC2+ 병합 Run: {merged_runs}건, 삭제 RC: {deleted_rc}개")
    print(f"  RELEASE-* legacy 처리: {legacy}건")
    print(f"  result 수: {before_results} → {after_results}")
    print(f"\n[타임라인 최상위 {len(top_levels)}행]")
    for r in top_levels:
        print(f"  {r[0]}  seq={r[1]}  {r[2]}")

    ok = before_results == after_results
    print(f"\n{'[OK]' if ok else '[FAIL]'} 검증  backup={backup_path}")
    conn.close()


if __name__ == "__main__":
    main()

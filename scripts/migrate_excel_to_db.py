#!/usr/bin/env python3
"""
Excel Test 관리 이력 → product_test_tracing_system DB 마이그레이션 스크립트
================================================================================
실행 방법 (Windows, 프로젝트 루트에서):
  python scripts/migrate_excel_to_db.py
  python scripts/migrate_excel_to_db.py "C:\\path\\to\\file.xlsx"

의존성:
  - Python 3.8+
  - openpyxl  (pip install openpyxl)

데이터 누락 방지 원칙:
  ① Results 참조 누락 Case ID 4개   → DRAFT skeleton 레코드 생성
  ② Results 참조 누락 Config ID 1개  → skeleton env 레코드 생성
  ③ HALF PASSED 45건 → BLOCKED (remark에 원본값 보존)
  ④ PASSED(XO/OO)   → PASSED  (remark에 원본값 보존)
  ⑤ IN TESTING      → IN_PROGRESS
  ⑥ 이슈 ID 다중값(개행 구분) → 각각 별도 defect 레코드
  ⑦ Run 단위 = (base_report_id, config_id) 조합
================================================================================
"""
from __future__ import annotations

import re
import sqlite3
import sys
import uuid
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

if len(sys.argv) > 1:
    EXCEL_PATH = Path(sys.argv[1])
else:
    EXCEL_PATH = PROJECT_ROOT.parent / "ai_coworking" / "test_tracing_system_to_migrate.xlsx"

PROJECT_ID  = "WIFI_CONNECTIVITY_TEST_2026"
MGRBY       = "migration_script_v1"
NOW         = datetime.now(timezone.utc).isoformat()

MAIN_DB_PATH = PROJECT_ROOT / "data" / "product_test_tracking_system.db"

# ── 상태값 매핑 ────────────────────────────────────────────────────────────────
RESULT_STATUS: dict[str, str] = {
    "PASSED":      "PASSED",
    "PASSED(XO)":  "PASSED",
    "PASSED(OO)":  "PASSED",
    "BLOCKED":     "BLOCKED",
    "HALF PASSED": "BLOCKED",
    "IN TESTING":  "IN_PROGRESS",
    "FAILED":      "FAILED",
}
REPORT_STATUS: dict[str, str] = {
    "PASSED":     "PASSED",
    "BLOCKED":    "BLOCKED",
    "IN TESTING": "IN_PROGRESS",
}

# ── DB 스키마 DDL ──────────────────────────────────────────────────────────────
SCHEMA_DDL = """
PRAGMA foreign_keys = OFF;

CREATE TABLE IF NOT EXISTS project (
    project_id          TEXT PRIMARY KEY,
    project_name        TEXT NOT NULL,
    project_description TEXT,
    project_status      TEXT NOT NULL DEFAULT 'ACTIVE',
    created_at          TEXT NOT NULL,
    created_by          TEXT NOT NULL,
    updated_at          TEXT NOT NULL,
    updated_by          TEXT NOT NULL,
    remark              TEXT
);

CREATE TABLE IF NOT EXISTS user_account (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    user_name     TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    role_name     TEXT NOT NULL,
    display_name  TEXT,
    phone_number  TEXT UNIQUE,
    company_name  TEXT,
    department_name TEXT,
    internal_title  TEXT,
    is_approved   INTEGER NOT NULL DEFAULT 1,
    created_at    TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at    TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS project_membership (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id      TEXT REFERENCES project(project_id),
    user_id         INTEGER NOT NULL REFERENCES user_account(id),
    membership_role TEXT NOT NULL DEFAULT 'tester',
    created_at      TEXT NOT NULL,
    created_by      TEXT NOT NULL,
    UNIQUE (project_id, user_id)
);

CREATE TABLE IF NOT EXISTS product_test_release (
    product_test_release_id     TEXT PRIMARY KEY,
    project_id                  TEXT,
    upstream_release_id         TEXT NOT NULL,
    upstream_release_system     TEXT NOT NULL,
    release_stage               TEXT NOT NULL,
    release_sequence            INTEGER NOT NULL,
    product_test_release_status TEXT NOT NULL,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_target_definition (
    product_test_target_definition_id     TEXT PRIMARY KEY,
    project_id                            TEXT,
    product_code                          TEXT NOT NULL,
    manufacturer                          TEXT NOT NULL,
    model_name                            TEXT NOT NULL,
    hardware_revision                     TEXT,
    default_software_version              TEXT,
    default_firmware_version              TEXT,
    product_test_target_definition_status TEXT NOT NULL,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_target (
    product_test_target_id            TEXT PRIMARY KEY,
    project_id                        TEXT,
    product_test_target_definition_id TEXT NOT NULL
        REFERENCES product_test_target_definition(product_test_target_definition_id),
    serial_number                     TEXT NOT NULL,
    software_version                  TEXT,
    firmware_version                  TEXT,
    manufacture_lot                   TEXT,
    product_test_target_status        TEXT NOT NULL,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_environment_definition (
    product_test_environment_definition_id     TEXT PRIMARY KEY,
    project_id                                 TEXT,
    product_test_environment_definition_name   TEXT NOT NULL,
    test_country       TEXT, test_city   TEXT, test_company TEXT,
    test_building      TEXT, test_floor  TEXT, test_room    TEXT,
    network_type       TEXT, test_computer_name TEXT,
    operating_system_version TEXT, test_tool_name TEXT, test_tool_version TEXT,
    power_voltage TEXT, power_frequency TEXT,
    power_connector_type TEXT, power_condition TEXT,
    product_test_environment_definition_status TEXT NOT NULL,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_environment (
    product_test_environment_id            TEXT PRIMARY KEY,
    project_id                             TEXT,
    product_test_environment_definition_id TEXT NOT NULL
        REFERENCES product_test_environment_definition(product_test_environment_definition_id),
    product_test_environment_name          TEXT NOT NULL,
    test_computer_name   TEXT, operating_system_version TEXT,
    test_tool_version    TEXT, network_type TEXT,
    power_voltage        TEXT, power_frequency TEXT,
    power_connector_type TEXT, captured_at TEXT,
    product_test_environment_status TEXT NOT NULL,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_case (
    product_test_case_id     TEXT PRIMARY KEY,
    project_id               TEXT,
    product_test_case_title  TEXT NOT NULL,
    test_category            TEXT NOT NULL,
    test_objective           TEXT,
    precondition             TEXT,
    expected_result          TEXT,
    product_test_case_status TEXT NOT NULL,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_procedure (
    product_test_procedure_id     TEXT PRIMARY KEY,
    project_id                    TEXT,
    product_test_case_id          TEXT NOT NULL
        REFERENCES product_test_case(product_test_case_id),
    procedure_sequence            INTEGER NOT NULL,
    procedure_action              TEXT NOT NULL,
    expected_result               TEXT,
    acceptance_criteria           TEXT NOT NULL,
    required_evidence_type        TEXT,
    product_test_procedure_status TEXT NOT NULL,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_run (
    product_test_run_id         TEXT PRIMARY KEY,
    project_id                  TEXT,
    product_test_release_id     TEXT NOT NULL
        REFERENCES product_test_release(product_test_release_id),
    product_test_target_id      TEXT NOT NULL
        REFERENCES product_test_target(product_test_target_id),
    product_test_environment_id TEXT NOT NULL
        REFERENCES product_test_environment(product_test_environment_id),
    product_test_run_status     TEXT NOT NULL,
    started_at    TEXT NOT NULL,
    started_by    TEXT NOT NULL,
    finished_at   TEXT,
    cancelled_at  TEXT,
    cancelled_by  TEXT,
    cancel_reason TEXT,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_report (
    product_test_report_id     TEXT PRIMARY KEY,
    project_id                 TEXT,
    product_test_release_id    TEXT NOT NULL
        REFERENCES product_test_release(product_test_release_id),
    product_test_report_type   TEXT NOT NULL,
    product_test_report_status TEXT NOT NULL,
    product_test_report_title  TEXT NOT NULL,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    approved_at TEXT, approved_by TEXT,
    rejected_at TEXT, rejected_by TEXT, rejection_reason TEXT,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_result (
    product_test_result_id     TEXT PRIMARY KEY,
    project_id                 TEXT,
    product_test_run_id        TEXT NOT NULL
        REFERENCES product_test_run(product_test_run_id),
    product_test_case_id       TEXT NOT NULL
        REFERENCES product_test_case(product_test_case_id),
    product_test_result_status TEXT NOT NULL,
    actual_result              TEXT,
    judgement_reason           TEXT,
    result_judged_at           TEXT,
    result_judged_by           TEXT,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_defect (
    product_test_defect_id     TEXT PRIMARY KEY,
    project_id                 TEXT,
    product_test_result_id     TEXT NOT NULL
        REFERENCES product_test_result(product_test_result_id),
    product_test_procedure_result_id TEXT,
    defect_title               TEXT NOT NULL,
    defect_description         TEXT NOT NULL,
    defect_severity            TEXT NOT NULL,
    defect_priority            TEXT NOT NULL,
    product_test_defect_status TEXT NOT NULL,
    assigned_to   TEXT, fixed_at TEXT, fixed_by TEXT, fix_description TEXT,
    retest_product_test_result_id TEXT,
    retested_at   TEXT, retested_by TEXT,
    closed_at     TEXT, closed_by  TEXT, rejection_reason TEXT,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_procedure_result (
    product_test_procedure_result_id TEXT PRIMARY KEY,
    project_id TEXT,
    product_test_result_id TEXT NOT NULL,
    product_test_procedure_id TEXT NOT NULL,
    product_test_procedure_result_status TEXT NOT NULL,
    actual_result TEXT, judgement_reason TEXT,
    judged_at TEXT, judged_by TEXT,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_evidence (
    product_test_evidence_id TEXT PRIMARY KEY,
    project_id TEXT,
    product_test_result_id TEXT NOT NULL,
    product_test_procedure_result_id TEXT,
    product_test_defect_id TEXT,
    product_test_evidence_type TEXT NOT NULL,
    file_name TEXT, file_path TEXT NOT NULL, file_hash TEXT,
    captured_at TEXT, captured_by TEXT,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    updated_at TEXT NOT NULL, updated_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_report_snapshot (
    product_test_report_snapshot_id TEXT PRIMARY KEY,
    project_id TEXT,
    product_test_report_id TEXT NOT NULL,
    product_test_release_id TEXT NOT NULL,
    snapshot_type TEXT NOT NULL,
    snapshot_format TEXT NOT NULL,
    snapshot_payload TEXT NOT NULL,
    snapshot_hash TEXT NOT NULL,
    source_data_locked INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    remark TEXT
);

CREATE TABLE IF NOT EXISTS product_test_status_transition (
    product_test_status_transition_id TEXT PRIMARY KEY,
    project_id TEXT,
    entity_type TEXT NOT NULL,
    entity_id TEXT NOT NULL,
    from_status TEXT,
    to_status TEXT NOT NULL,
    transition_reason TEXT,
    transitioned_at TEXT NOT NULL,
    transitioned_by TEXT NOT NULL,
    created_at TEXT NOT NULL, created_by TEXT NOT NULL,
    remark TEXT
);

PRAGMA foreign_keys = ON;
"""

# ── 유틸 ──────────────────────────────────────────────────────────────────────
def safe(v: object) -> str:
    return str(v).strip() if v is not None else ""


def map_result_status(raw: str) -> tuple[str, str | None]:
    mapped   = RESULT_STATUS.get(raw, raw)
    original = raw if mapped != raw else None
    return mapped, original


def parse_detail(detail: str) -> tuple[str, str, str]:
    s = detail or ""
    tm = re.search(r"검토자\s*:\s*([^,\n]+)", s)
    dm = re.search(r"시험일자\s*:\s*([^,\n]+)", s)
    sm = re.search(r"내용요약\s*:\s*(.+?)(?:\n|$)", s)
    return (
        tm.group(1).strip() if tm else "",
        dm.group(1).strip() if dm else "",
        sm.group(1).strip() if sm else s.strip(),
    )


def parse_issues(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [i.strip() for i in re.split(r"[\n\r,;]+", raw) if i.strip()]


def parse_targets(raw: str) -> list[dict]:
    out = []
    if not raw:
        return out
    m = re.search(r"\[Test 대상\](.*?)(?:\[|$)", raw, re.DOTALL)
    if not m:
        return out
    for line in m.group(1).split("\n"):
        line = line.strip()
        h = re.match(r"(H[A-Z]{2}-\d+[A-Z0-9]*)(?:\s+(\S+))?(?:\s+(.*))?", line)
        if h and h.group(1):
            rest   = (h.group(3) or "").strip()
            serial = rest.split("(")[0].strip() or "UNKNOWN"
            out.append({
                "model":   h.group(1).strip(),
                "sw":      h.group(2).strip() if h.group(2) else "UNKNOWN",
                "serial":  serial,
                "remark":  rest or None,
            })
    return out


def base_rid(full_id: str) -> str:
    return full_id.rsplit("_순번", 1)[0]


def slug(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "_", s)


# ── 연결 + 스키마 ──────────────────────────────────────────────────────────────
def open_db(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(path))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn


def exists(conn: sqlite3.Connection, table: str, pk_col: str, pk_val: str) -> bool:
    row = conn.execute(
        f"SELECT 1 FROM {table} WHERE {pk_col}=? LIMIT 1", (pk_val,)
    ).fetchone()
    return row is not None


# ── 메인 ──────────────────────────────────────────────────────────────────────
def main() -> None:
    try:
        import openpyxl
    except ImportError:
        sys.exit("[ERROR] openpyxl 없음. pip install openpyxl 실행 후 재시도")

    print("=" * 65)
    print("  Excel → DB 마이그레이션")
    print(f"  Excel  : {EXCEL_PATH}")
    print(f"  Main DB: {MAIN_DB_PATH}")
    print(f"  Project: {PROJECT_ID}")
    print("=" * 65)

    if not EXCEL_PATH.exists():
        sys.exit(f"[ERROR] Excel 없음: {EXCEL_PATH}")

    # ── Excel 파싱 ─────────────────────────────────────────────────────────────
    wb           = openpyxl.load_workbook(str(EXCEL_PATH))
    ws_sc        = wb["Test Scenarios"]
    ws_ca        = wb["Test Cases"]
    ws_co        = wb["Test Configs"]
    ws_rp        = wb["Test Reports"]
    ws_rs        = wb["Test Results"]

    case_steps: dict[str, list[dict]]  = defaultdict(list)
    case_title: dict[str, str]         = {}
    for row in ws_ca.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            continue
        cid = safe(row[1])
        case_steps[cid].append({
            "action":   safe(row[2]),
            "criteria": safe(row[3]),
            "remark":   safe(row[4]) or None,
        })
        if cid not in case_title and row[3]:
            case_title[cid] = safe(row[3])[:120]

    scenario_map: dict[str, list[dict]] = defaultdict(list)
    for row in ws_sc.iter_rows(min_row=3, values_only=True):
        cid = safe(row[4]) if row[4] else None
        if not cid:
            continue
        scenario_map[cid].append({
            "pub":   safe(row[1]).replace("\n", "").strip(),
            "int":   safe(row[2]).replace("\n", "").strip(),
            "title": safe(row[3]),
        })

    configs: list[dict] = []
    for row in ws_co.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            continue
        configs.append({
            "id":     safe(row[1]),
            "alias":  safe(row[2]),
            "purp":   safe(row[3]),
            "tgt":    safe(row[4]),
            "env":    safe(row[5]),
            "stat":   safe(row[6]) or "사용 중",
            "chg":    safe(row[7]),
            "remark": safe(row[8]) or None,
        })

    reports: list[dict] = []
    for row in ws_rp.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        reports.append({
            "no":       row[0],
            "alias":    safe(row[1]),
            "obj_raw":  safe(row[2]),
            "stat_raw": safe(row[3]) or "IN TESTING",
            "remark":   safe(row[4]) or None,
            "rids_raw": safe(row[5]),
            "workday":  safe(row[6]) or None,
            "start":    safe(row[7]) or None,
            "end":      safe(row[8]) or None,
        })

    results: list[dict] = []
    for row in ws_rs.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        results.append({
            "no":      row[0],
            "rid_f":   safe(row[1]),
            "combo":   safe(row[2]),
            "cfg_id":  safe(row[3]),
            "case_id": safe(row[4]),
            "stat_r":  safe(row[5]),
            "detail":  safe(row[6]),
            "issue_r": safe(row[7]) or None,
            "remark":  safe(row[8]) or None,
        })

    # ── 누락 확인 ─────────────────────────────────────────────────────────────
    ex_cfg  = {c["id"] for c in configs}
    ex_case = set(case_steps.keys())
    miss_cfg  = {r["cfg_id"] for r in results if r["cfg_id"]} - ex_cfg
    miss_case = {r["case_id"] for r in results if r["case_id"]} - ex_case

    print(f"\n[사전검증] 누락 Config {len(miss_cfg)}건 → skeleton 처리")
    for c in sorted(miss_cfg):
        print(f"    ⚠  {c}")
    print(f"[사전검증] 누락 Case  {len(miss_case)}건 → skeleton 처리")
    for c in sorted(miss_case):
        print(f"    ⚠  {c}")
    print(f"[사전검증] Excel Result 총 {len(results)}건 → 전량 삽입\n")

    # ── DB 초기화 ──────────────────────────────────────────────────────────────
    conn = open_db(MAIN_DB_PATH)
    conn.executescript(SCHEMA_DDL)
    conn.commit()
    print("[OK] DB 스키마 준비 완료")

    # ── STEP 1: project ────────────────────────────────────────────────────────
    if not exists(conn, "project", "project_id", PROJECT_ID):
        conn.execute("""
            INSERT INTO project
              (project_id,project_name,project_description,project_status,
               created_at,created_by,updated_at,updated_by,remark)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            PROJECT_ID,
            "Wi-Fi 연결성 시험 프로젝트 2026",
            "제품 Wi-Fi 연결성 시험 이력 마이그레이션 (Excel → DB)",
            "ACTIVE", NOW, MGRBY, NOW, MGRBY,
            f"마이그레이션 일시: {NOW}",
        ))
    conn.commit()
    print("  [1/8] project 완료")

    # ── STEP 2: user_account + project_membership ──────────────────────────────
    tester_names: set[str] = set()
    for r in results:
        t, _, _ = parse_detail(r["detail"])
        if t:
            tester_names.add(t)

    def upsert_user(uname: str, display: str, role: str) -> None:
        if not exists(conn, "user_account", "user_name", uname):
            conn.execute("""
                INSERT INTO user_account
                  (user_name,password_hash,role_name,display_name,
                   is_approved,created_at,updated_at)
                VALUES (?,?,?,?,1,?,?)
            """, (uname, uname, role, display, NOW, NOW))

    upsert_user(MGRBY, "마이그레이션 스크립트", "master_admin")
    for name in sorted(tester_names):
        uname = re.sub(r"\s+", "_", name.strip()).lower()
        upsert_user(uname, name, "admin")

    conn.commit()

    # membership: migration_script 계정
    mgr_uid = conn.execute(
        "SELECT id FROM user_account WHERE user_name=?", (MGRBY,)
    ).fetchone()[0]
    if not conn.execute(
        "SELECT 1 FROM project_membership WHERE project_id=? AND user_id=?",
        (PROJECT_ID, mgr_uid)
    ).fetchone():
        conn.execute("""
            INSERT INTO project_membership
              (project_id,user_id,membership_role,created_at,created_by)
            VALUES (?,?,?,?,?)
        """, (PROJECT_ID, mgr_uid, "master_admin", NOW, MGRBY))
    conn.commit()
    print(f"  [2/8] user_account {len(tester_names)+1}건 완료")

    # ── STEP 3: target_definition + target ─────────────────────────────────────
    model_def: dict[str, str] = {}         # model → def_id
    target_id_map: dict[tuple, str] = {}   # (model, serial) → target_id

    all_tgt_lines: list[dict] = []
    for rpt in reports:
        all_tgt_lines.extend(parse_targets(rpt["obj_raw"]))

    seen: set[tuple] = set()
    uniq_tgts = []
    for t in all_tgt_lines:
        k = (t["model"], t["serial"])
        if k not in seen:
            seen.add(k)
            uniq_tgts.append(t)

    for t in uniq_tgts:
        model, serial = t["model"], t["serial"]
        def_id = f"TARGET_DEF-{model.replace('-','_')}"
        if model not in model_def:
            if not exists(conn, "product_test_target_definition",
                          "product_test_target_definition_id", def_id):
                conn.execute("""
                    INSERT INTO product_test_target_definition
                      (product_test_target_definition_id,project_id,product_code,
                       manufacturer,model_name,default_software_version,
                       product_test_target_definition_status,
                       created_at,created_by,updated_at,updated_by)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, (def_id, PROJECT_ID, model, "Huvitz", model, t["sw"],
                      "ACTIVE", NOW, MGRBY, NOW, MGRBY))
            model_def[model] = def_id

        tgt_id = f"TARGET-{slug(model)}-{slug(serial)}"
        if not exists(conn, "product_test_target",
                      "product_test_target_id", tgt_id):
            conn.execute("""
                INSERT INTO product_test_target
                  (product_test_target_id,project_id,
                   product_test_target_definition_id,
                   serial_number,software_version,
                   product_test_target_status,
                   created_at,created_by,updated_at,updated_by,remark)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (tgt_id, PROJECT_ID, model_def[model], serial, t["sw"],
                  "ACTIVE", NOW, MGRBY, NOW, MGRBY, t.get("remark")))
        target_id_map[(model, serial)] = tgt_id

    conn.commit()

    # report별 primary target
    rpt_primary_tgt: dict[str, str] = {}
    for rpt in reports:
        lines = parse_targets(rpt["obj_raw"])
        if lines:
            k = (lines[0]["model"], lines[0]["serial"])
            rpt_primary_tgt[rpt["alias"]] = target_id_map.get(k, "")
    fallback_tgt = list(target_id_map.values())[0] if target_id_map else ""

    print(f"  [3/8] target_def {len(model_def)}건, target {len(target_id_map)}건 완료")

    # ── STEP 4: env_definition + environment ───────────────────────────────────
    env_def_map: dict[str, str] = {}   # config_id → env_def_id
    env_map:     dict[str, str] = {}   # config_id → env_id

    def upsert_env(cid: str, alias: str, purp: str, tgt_d: str,
                   env_d: str, chg: str, remark_extra: str | None,
                   skeleton: bool = False) -> None:
        ed_id = f"ENV_DEF-{cid}"
        e_id  = f"ENV-{cid}"
        note  = ("[마이그레이션 주의] 원본 Configs 시트에 정의 없음. Results에서 참조됨.\n"
                 if skeleton else "")
        full_remark = (
            f"{note}[Config ID] {cid}\n"
            f"[목적] {purp[:200]}\n"
            f"[시험대상] {tgt_d[:300]}\n"
            f"[환경] {env_d[:400]}\n"
            f"[변경점] {chg[:200]}"
        )
        if not exists(conn, "product_test_environment_definition",
                      "product_test_environment_definition_id", ed_id):
            conn.execute("""
                INSERT INTO product_test_environment_definition
                  (product_test_environment_definition_id,project_id,
                   product_test_environment_definition_name,
                   product_test_environment_definition_status,
                   created_at,created_by,updated_at,updated_by,remark)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (ed_id, PROJECT_ID, alias or cid, "ACTIVE",
                  NOW, MGRBY, NOW, MGRBY, full_remark))

        if not exists(conn, "product_test_environment",
                      "product_test_environment_id", e_id):
            conn.execute("""
                INSERT INTO product_test_environment
                  (product_test_environment_id,project_id,
                   product_test_environment_definition_id,
                   product_test_environment_name,
                   product_test_environment_status,
                   created_at,created_by,updated_at,updated_by,remark)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (e_id, PROJECT_ID, ed_id, alias or cid, "ACTIVE",
                  NOW, MGRBY, NOW, MGRBY, remark_extra))

        env_def_map[cid] = ed_id
        env_map[cid]     = e_id

    for c in configs:
        upsert_env(c["id"], c["alias"], c["purp"], c["tgt"],
                   c["env"], c["chg"], c["remark"], skeleton=False)
    for cid in sorted(miss_cfg):
        upsert_env(cid, f"[SKELETON] {cid}",
                   "누락된 Config ID — skeleton 레코드 (마이그레이션 보완)",
                   "", "", "", None, skeleton=True)

    conn.commit()
    fallback_env = list(env_map.values())[0] if env_map else ""
    print(f"  [4/8] env_def {len(env_def_map)}건, env {len(env_map)}건 완료")

    # ── STEP 5: product_test_release ───────────────────────────────────────────
    release_map: dict[str, str] = {}   # base_rid → release_id

    for rpt in reports:
        raw = rpt["rids_raw"]
        if not raw or raw.upper() == "TBD":
            continue
        rstatus = REPORT_STATUS.get(rpt["stat_raw"], rpt["stat_raw"])
        seq     = int(rpt["no"]) if isinstance(rpt["no"], int) else 1
        for rid in raw.split("\n"):
            rid = rid.strip()
            if not rid:
                continue
            rel_id = f"RELEASE-{rid}"
            if not exists(conn, "product_test_release",
                          "product_test_release_id", rel_id):
                conn.execute("""
                    INSERT INTO product_test_release
                      (product_test_release_id,project_id,
                       upstream_release_id,upstream_release_system,
                       release_stage,release_sequence,
                       product_test_release_status,
                       created_at,created_by,updated_at,updated_by,remark)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, (rel_id, PROJECT_ID, rid, "EXCEL_MIGRATION",
                      "TEST", seq, rstatus, NOW, MGRBY, NOW, MGRBY,
                      f"[Report Alias] {rpt['alias']}\n"
                      f"[Workday] {rpt['workday']}\n"
                      f"[Start] {rpt['start']}  [End] {rpt['end']}"))
            release_map[rid] = rel_id

    FALLBACK_REL = f"RELEASE-FALLBACK-{PROJECT_ID}"
    if not exists(conn, "product_test_release",
                  "product_test_release_id", FALLBACK_REL):
        conn.execute("""
            INSERT INTO product_test_release
              (product_test_release_id,project_id,
               upstream_release_id,upstream_release_system,
               release_stage,release_sequence,
               product_test_release_status,
               created_at,created_by,updated_at,updated_by,remark)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (FALLBACK_REL, PROJECT_ID, "MIGRATION_FALLBACK",
              "EXCEL_MIGRATION", "TEST", 99, "IN_PROGRESS",
              NOW, MGRBY, NOW, MGRBY,
              "TBD Report IDs가 있는 시험 Report를 위한 fallback release"))
    conn.commit()
    print(f"  [5/8] release {len(release_map)}건 + fallback 완료")

    # ── STEP 6: test_case + procedure ──────────────────────────────────────────
    reg_cases: set[str] = set()

    def upsert_case(cid: str, skeleton: bool = False) -> None:
        if cid in reg_cases:
            return
        scs  = scenario_map.get(cid, [])
        note = ""
        if scs:
            note = "[연결 시나리오]\n" + "\n".join(
                f"  · {s['pub']} / {s['int']} : {s['title']}" for s in scs
            )
        title = case_title.get(cid, cid)
        if not exists(conn, "product_test_case",
                      "product_test_case_id", cid):
            conn.execute("""
                INSERT INTO product_test_case
                  (product_test_case_id,project_id,
                   product_test_case_title,test_category,
                   product_test_case_status,
                   created_at,created_by,updated_at,updated_by,remark)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (cid, PROJECT_ID,
                  ("[SKELETON] " if skeleton else "") + title,
                  "WIFI_CONNECTIVITY",
                  "DRAFT" if skeleton else "ACTIVE",
                  NOW, MGRBY, NOW, MGRBY,
                  ("[마이그레이션 주의] Cases 시트에 정의 없음. Results에서 참조됨.\n"
                   if skeleton else "") + note))

        steps = case_steps.get(cid, [])
        for seq, step in enumerate(steps, start=1):
            proc_id = f"{cid}_STEP_{seq:03d}"
            if not exists(conn, "product_test_procedure",
                          "product_test_procedure_id", proc_id):
                conn.execute("""
                    INSERT INTO product_test_procedure
                      (product_test_procedure_id,project_id,
                       product_test_case_id,procedure_sequence,
                       procedure_action,acceptance_criteria,
                       product_test_procedure_status,
                       created_at,created_by,updated_at,updated_by,remark)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, (proc_id, PROJECT_ID, cid, seq,
                      step["action"], step["criteria"], "ACTIVE",
                      NOW, MGRBY, NOW, MGRBY, step["remark"]))
        reg_cases.add(cid)

    for cid in sorted(case_steps.keys()):
        upsert_case(cid, skeleton=False)
    for cid in sorted(miss_case):
        upsert_case(cid, skeleton=True)

    conn.commit()
    proc_cnt = conn.execute(
        "SELECT COUNT(*) FROM product_test_procedure WHERE project_id=?",
        (PROJECT_ID,)
    ).fetchone()[0]
    print(f"  [6/8] test_case {len(reg_cases)}건, procedure {proc_cnt}건 완료")

    # ── STEP 7: test_report + test_run ─────────────────────────────────────────
    report_id_map: dict[str, str] = {}     # base_rid → db_report_id
    run_id_map:    dict[tuple, str] = {}   # (base_rid, cfg_id) → run_id

    for rpt in reports:
        raw = rpt["rids_raw"]
        base_rids = []
        if raw and raw.upper() != "TBD":
            base_rids = [r.strip() for r in raw.split("\n") if r.strip()]
        if not base_rids:
            base_rids = [f"TBD_REPORT_NO{rpt['no']}"]

        rstatus  = REPORT_STATUS.get(rpt["stat_raw"], rpt["stat_raw"])
        prim_tgt = rpt_primary_tgt.get(rpt["alias"], fallback_tgt)
        started  = (f"{rpt['start']}T00:00:00+00:00" if rpt["start"] else NOW)
        ended    = (f"{rpt['end']}T23:59:59+00:00"   if rpt["end"]   else None)

        for rid in base_rids:
            rel_id      = release_map.get(rid, FALLBACK_REL)
            db_rpt_id   = f"RPT-{rid}"

            if not exists(conn, "product_test_report",
                          "product_test_report_id", db_rpt_id):
                conn.execute("""
                    INSERT INTO product_test_report
                      (product_test_report_id,project_id,
                       product_test_release_id,
                       product_test_report_type,
                       product_test_report_status,
                       product_test_report_title,
                       created_at,created_by,updated_at,updated_by,remark)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, (db_rpt_id, PROJECT_ID, rel_id,
                      "TEST_REPORT", rstatus, rpt["alias"],
                      NOW, MGRBY, NOW, MGRBY, rpt["remark"]))
            report_id_map[rid] = db_rpt_id

            # run: (rid, config_id) 조합
            cfgs_used = sorted({
                r["cfg_id"]
                for r in results
                if base_rid(r["rid_f"]) == rid and r["cfg_id"]
            })
            if not cfgs_used:
                cfgs_used = [list(env_map.keys())[0]] if env_map else ["UNKNOWN"]

            for cfg_id in cfgs_used:
                env_id   = env_map.get(cfg_id, fallback_env)
                run_id   = f"RUN-{rid}-CFG_{slug(cfg_id)}"
                run_key  = (rid, cfg_id)
                if not exists(conn, "product_test_run",
                              "product_test_run_id", run_id):
                    conn.execute("""
                        INSERT INTO product_test_run
                          (product_test_run_id,project_id,
                           product_test_release_id,
                           product_test_target_id,
                           product_test_environment_id,
                           product_test_run_status,
                           started_at,started_by,finished_at,
                           created_at,created_by,updated_at,updated_by,remark)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (run_id, PROJECT_ID, rel_id,
                          prim_tgt or fallback_tgt, env_id,
                          rstatus, started, MGRBY, ended,
                          NOW, MGRBY, NOW, MGRBY,
                          f"[Workday] {rpt['workday']}\n"
                          f"[Config] {cfg_id}\n"
                          f"[Test Object]\n{rpt['obj_raw'][:400]}"))
                run_id_map[run_key] = run_id

    conn.commit()
    run_cnt = conn.execute(
        "SELECT COUNT(*) FROM product_test_run WHERE project_id=?",
        (PROJECT_ID,)
    ).fetchone()[0]
    print(f"  [7/8] test_report {len(report_id_map)}건, test_run {run_cnt}건 완료")

    # ── STEP 8: test_result + test_defect ──────────────────────────────────────
    res_cnt    = 0
    defect_cnt = 0
    warn_cnt   = 0

    for r in results:
        brid   = base_rid(r["rid_f"])
        cfg_id = r["cfg_id"]
        cid    = r["case_id"]
        rkey   = (brid, cfg_id)

        # run_id 결정
        run_id = run_id_map.get(rkey)
        if not run_id:
            candidates = [v for k, v in run_id_map.items() if k[0] == brid]
            if candidates:
                run_id = candidates[0]
            else:
                # base_rid 자체가 없는 경우 → fallback run 생성
                fallback_run_id = f"RUN-FALLBACK-{slug(brid)}"
                if not exists(conn, "product_test_run",
                              "product_test_run_id", fallback_run_id):
                    conn.execute("""
                        INSERT INTO product_test_run
                          (product_test_run_id,project_id,
                           product_test_release_id,
                           product_test_target_id,
                           product_test_environment_id,
                           product_test_run_status,
                           started_at,started_by,
                           created_at,created_by,updated_at,updated_by,remark)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (fallback_run_id, PROJECT_ID, FALLBACK_REL,
                          fallback_tgt, fallback_env, "IN_PROGRESS",
                          NOW, MGRBY, NOW, MGRBY, NOW, MGRBY,
                          f"[마이그레이션 주의] Report ID '{brid}'에 대응하는 Run 없음 → fallback"))
                run_id_map[rkey] = fallback_run_id
                run_id = fallback_run_id
        # case_id 없는 미완성 행 → PLACEHOLDER case에 연결해서 데이터 보존
        if not cid:
            cid = f"PLACEHOLDER_EMPTY_CASE-{PROJECT_ID}"
            if cid not in reg_cases:
                if not exists(conn, "product_test_case",
                              "product_test_case_id", cid):
                    conn.execute("""
                        INSERT INTO product_test_case
                          (product_test_case_id,project_id,
                           product_test_case_title,test_category,
                           product_test_case_status,
                           created_at,created_by,updated_at,updated_by,remark)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, (cid, PROJECT_ID,
                          "[PLACEHOLDER] Case ID 미입력 결과 행 보존용",
                          "WIFI_CONNECTIVITY", "DRAFT",
                          NOW, MGRBY, NOW, MGRBY,
                          "[마이그레이션 주의] Excel Results 시트에서 Case ID가 비어 있는 행들을 "
                          "데이터 누락 없이 보존하기 위한 placeholder case"))
                reg_cases.add(cid)

        if not run_id or cid not in reg_cases:
            warn_cnt += 1
            print(f"    [WARN] no={r['no']} — "
                  f"run={run_id or 'NONE'}, case={cid or 'NONE'} 건너뜀")
            continue

        mapped_st, orig_st = map_result_status(r["stat_r"])
        tester, tested_dt, summary = parse_detail(r["detail"])
        result_id = f"RESULT-{brid}-NO{r['no']:05d}"

        remarks = []
        if orig_st:
            remarks.append(f"[원본상태] {orig_st}")
        if r["combo"]:
            remarks.append(f"[연결구성] {r['combo']}")
        if r["remark"]:
            remarks.append(f"[비고] {r['remark']}")

        if not exists(conn, "product_test_result",
                      "product_test_result_id", result_id):
            conn.execute("""
                INSERT INTO product_test_result
                  (product_test_result_id,project_id,
                   product_test_run_id,product_test_case_id,
                   product_test_result_status,
                   actual_result,result_judged_at,result_judged_by,
                   created_at,created_by,updated_at,updated_by,remark)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (result_id, PROJECT_ID, run_id, cid,
                  mapped_st, summary or None,
                  (f"{tested_dt}T00:00:00+00:00" if tested_dt else NOW),
                  tester or MGRBY,
                  NOW, MGRBY, NOW, MGRBY,
                  "\n".join(remarks) or None))
            res_cnt += 1

            for idx, issue_id in enumerate(parse_issues(r["issue_r"]), start=1):
                def_id = f"DEFECT-{slug(issue_id)}-RNO{r['no']:05d}-{idx:02d}"
                if not exists(conn, "product_test_defect",
                              "product_test_defect_id", def_id):
                    conn.execute("""
                        INSERT INTO product_test_defect
                          (product_test_defect_id,project_id,
                           product_test_result_id,
                           defect_title,defect_description,
                           defect_severity,defect_priority,
                           product_test_defect_status,
                           created_at,created_by,updated_at,updated_by)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (def_id, PROJECT_ID, result_id,
                          issue_id,
                          f"[이슈ID] {issue_id}\n"
                          f"[시험결과] {r['stat_r']}\n"
                          f"[연결구성] {r['combo']}\n"
                          f"[요약] {summary}",
                          "MEDIUM", "MEDIUM", "OPEN",
                          NOW, MGRBY, NOW, MGRBY))
                    defect_cnt += 1

    conn.commit()
    print(f"  [8/8] test_result {res_cnt}건, defect {defect_cnt}건 완료"
          + (f" (건너뜀 {warn_cnt}건)" if warn_cnt else ""))

    # ── 최종 검증 ─────────────────────────────────────────────────────────────
    print("\n" + "=" * 65)
    print("  최종 검증")
    print("=" * 65)

    def cnt(table: str, col: str = "project_id") -> int:
        return conn.execute(
            f"SELECT COUNT(*) FROM {table} WHERE {col}=?", (PROJECT_ID,)
        ).fetchone()[0]

    rows = [
        ("project",                      "project",                    "project_id"),
        ("user_account",                 "user_account",               None),
        ("product_test_target_def",      "product_test_target_definition", "project_id"),
        ("product_test_target",          "product_test_target",        "project_id"),
        ("env_definition",               "product_test_environment_definition", "project_id"),
        ("environment",                  "product_test_environment",   "project_id"),
        ("release",                      "product_test_release",       "project_id"),
        ("test_case",                    "product_test_case",          "project_id"),
        ("procedure",                    "product_test_procedure",     "project_id"),
        ("test_report",                  "product_test_report",        "project_id"),
        ("test_run",                     "product_test_run",           "project_id"),
        ("test_result",                  "product_test_result",        "project_id"),
        ("test_defect",                  "product_test_defect",        "project_id"),
    ]
    for label, table, col in rows:
        if col:
            c = conn.execute(
                f"SELECT COUNT(*) FROM {table} WHERE {col}=?", (PROJECT_ID,)
            ).fetchone()[0]
        else:
            c = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  {label:<35s}: {c:>5d}건")

    db_res     = conn.execute(
        "SELECT COUNT(*) FROM product_test_result WHERE project_id=?",
        (PROJECT_ID,)
    ).fetchone()[0]
    excel_total = len(results)
    ok  = (db_res == excel_total)
    sym = "✅" if ok else "⚠ "
    print()
    print(f"  {sym} Result 일치: DB {db_res}건 / Excel {excel_total}건"
          + (" → 완전 일치" if ok else f" → 차이 {excel_total - db_res}건 ← 확인 필요"))

    sk_case = conn.execute(
        "SELECT COUNT(*) FROM product_test_case "
        "WHERE project_id=? AND product_test_case_status='DRAFT'",
        (PROJECT_ID,)
    ).fetchone()[0]
    sk_env = conn.execute(
        "SELECT COUNT(*) FROM product_test_environment_definition "
        "WHERE project_id=? AND "
        "product_test_environment_definition_name LIKE '[SKELETON]%'",
        (PROJECT_ID,)
    ).fetchone()[0]
    if sk_case:
        print(f"  ℹ  skeleton test_case (DRAFT) {sk_case}건 → 내용 보완 필요")
    if sk_env:
        print(f"  ℹ  skeleton env_definition     {sk_env}건 → Config 정보 보완 필요")

    conn.close()
    print("\n[완료] 마이그레이션 성공")

if __name__ == "__main__":
    main()

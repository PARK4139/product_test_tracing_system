from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.services.product_test_run_service import list_runs


def build_test_result_workbook(database_session: Session) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "product_test_runs"
    worksheet.append(
        [
            "product_test_run_id",
            "test_round_id",
            "product_test_target_id",
            "product_test_config_id",
            "product_test_run_status",
            "started_at",
            "started_by",
        ]
    )
    for row in list_runs(database_session):
        worksheet.append(
            [
                row.get("product_test_run_id"),
                row.get("test_round_id"),
                row.get("product_test_target_id"),
                row.get("product_test_config_id"),
                row.get("product_test_run_status"),
                row.get("started_at"),
                row.get("started_by"),
            ]
        )
    return workbook


def append_test_results_to_existing_workbook(
    *,
    database_session: Session,
    excel_file_path: str,
    sheet_name: str,
    limit: int = 1000,
) -> dict[str, int | str]:
    path = (excel_file_path or "").strip()
    if not path:
        raise ValueError("excel_file_path is required.")
    target_sheet = (sheet_name or "").strip()
    if not target_sheet:
        raise ValueError("sheet_name is required.")

    workbook = load_workbook(path)
    worksheet = workbook[target_sheet] if target_sheet in workbook.sheetnames else workbook.create_sheet(target_sheet)

    if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
        worksheet.append(["product_test_run_id", "test_round_id", "product_test_run_status"])

    appended = 0
    for row in list_runs(database_session)[:limit]:
        worksheet.append(
            [
                row.get("product_test_run_id"),
                row.get("test_round_id"),
                row.get("product_test_run_status"),
            ]
        )
        appended += 1

    workbook.save(path)
    return {"sheet_name": target_sheet, "appended_rows": appended}
